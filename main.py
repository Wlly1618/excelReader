import openpyxl
import csv
import json


def get_data(worksheet, max_r: int, min_r: int, max_c: int, min_c: int):
    data_row = []
    for row in worksheet.iter_rows(max_row=max_r, max_col=max_c, min_row=min_r, min_col=min_c, values_only=True):
        data_row.append(row)
    return data_row


def writer_file_out(name_file: str, type_file: int, header_write, data_write):
    if type_file == 1:
        print("File type txt")
        with open(name_file + '.txt', 'w') as file_txt:
            for header in header_write:
                file_txt.write(header + "\t")
            file_txt.write("\n")
            for data_row in data_write:
                for data in data_row:
                    print(data)
                    file_txt.write(str(data))
                    file_txt.write("\t")
                file_txt.write("\n")
            file_txt.close()

    if type_file == 2:
        print("File type csv")
        with open(name_file + '.csv', 'w', newline='') as file_csv:
            writer = csv.writer(file_csv)
            writer.writerow(header_write)
            for row in data_write:
                writer.writerow(row)
            file_csv.close()

    if type_file == 3:
        print("File type JSON")
        with open(name_file + ".json", 'w', newline='') as file_json:
            for values in data_write:
                dictionary = dict(zip(header_write, values))
                obj = type("DATA", (object,), dictionary)()
                


def main():
    route = input("Enter route to excel: ")
    workbook = openpyxl.load_workbook(route)

    position = int(input("Enter position to sheet to read data: "))
    sheet = workbook.sheetnames[position]
    worksheet = workbook[sheet]

    row_header = input("Where is the Header: ")
    header = worksheet[row_header]

    header_values = []
    for head in header:
        header_values.append(head.value)

    values = input("Where is the data? [Min row] [Max row] [Min Col] [Max Col] : ")
    min_row, max_row, min_col, max_col = map(int, values.split())
    data = get_data(worksheet=worksheet, max_r=max_row, min_r=min_row, max_c=max_col, min_c=min_col)

    name_file = input("Name file output: ")
    type_file = int(input("Type file [ 1 .txt ] [ 2 .csv ] [3 .json]: "))

    writer_file_out(name_file=name_file, type_file=type_file, header_write=header_values, data_write=data)


if __name__ == '__main__':
    main()
